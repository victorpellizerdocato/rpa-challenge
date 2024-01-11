import os
import random
import logging
import requests
from string import digits
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from dateutil.relativedelta import relativedelta


class DataHandling:
    def get_last_acceptable_date(
        months_delta: int
    ) -> datetime:
        logging.info("Defining the last month to look for news.")
        if months_delta <= 1:
            last_acceptable_date = datetime.utcnow().replace(day=1, hour=0)
        else:
            last_acceptable_date = datetime.today() - \
                relativedelta(months=months_delta-1)
        return last_acceptable_date

    def date_filter(
        date: str
    ) -> datetime:
        logging.info("Converting the news' date format to datetime format.")
        months = {
            'Jan': 1,
            'Feb': 2,
            'March': 3,
            'April': 4,
            'May': 5,
            'June': 6,
            'July': 7,
            'Aug': 8,
            'Sept': 9,
            'Oct': 10,
            'Nov': 11,
            'Dec': 12
        }
        if 'hour' in date:
            return datetime.today()
        split_date = date.split(' ')
        # the month sometimes is misspelled
        month = months[split_date[0].replace('.', '')]
        day = split_date[1].replace(',', '')
        filtered_date = datetime.strptime(
            f"{day}/{month}/{split_date[2]}", "%d/%m/%Y")
        return filtered_date

    def download_file(
        url: str,
        date: str,
        query: str
    ) -> str:
        download_file_response = {
            'success': False
        }
        try:
            logging.info("Downloading the new's image.")
            download_response = requests.get(
                url=url
            )
            if download_response and download_response.status_code != 200:
                return ''

            file_name = f'{query}-{date}-{random.randint(1000,9999)}.png'
            path = f'./output/{file_name}'
            open(path, "wb").write(download_response.content)
            file_size = os.path.getsize(path)
            if file_size:
                logging.info("Successfully downloaded the new's image.")

            download_file_response.update({
                'success': True,
                'file_size': file_size,
                'picture_filename': file_name
            })

        finally:
            return download_file_response

    def build_sheet(
        extracted_data: list,
        sheet_name: str
    ) -> str:
        logging.info("Building sheet with the extracted data.")
        header = [
            'picture_filename',
            'title',
            'description',
            'date',
            'search_phrase_count',
            'contains_money'
        ]
        header_len = len(header)
        header_cells = []

        for index in range(header_len):
            header_cells.append(f'{chr(index+65)}1')
        cell_number = str.maketrans('', '', digits)

        logging.info(f'Extracted data: {extracted_data}')
        wb = Workbook()
        ws = wb.active

        fill_style = PatternFill(
            start_color='205792',
            end_color='205792',
            fill_type='solid'
        )

        font_style = Font(
            bold=True,
            size=12,
            color='FFFFFF',
            name='Calibri'
        )

        for index, cell in enumerate(header_cells):
            column = cell.translate(cell_number)
            ws[cell].fill = fill_style
            ws[cell].font = font_style
            ws[cell] = header[index]
            ws.column_dimensions[column].width = len(header[index]) + 25

            for index_row, process in enumerate(extracted_data):
                ws.cell(
                    column=index+1,
                    row=index_row+2,
                    value=process[header[index]]
                )

        sheet_path = f'./output/{sheet_name}.xlsx'
        wb.save(sheet_path)
        return sheet_path
