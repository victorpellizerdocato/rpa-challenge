"""

> Packages used to handle general data

logging - logs every relevant step
os - path handling
string - supplies the alphabet and digits
uuid - generates a unique string
datetime - multiple date operations
quote - changes whitespaces from queries
relativedelta - makes date comparison operations easier
Workbook - represents the entire spreadsheet
Font - define cell's font
PatternFill - define cell's color
Selenium - the browser used to scrape the web

"""
import logging
import os
import string
import uuid
from datetime import datetime
from urllib.parse import quote

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from RPA.Browser.Selenium import Selenium


class DataHandling():
    """
    Class where general data handling methods are placed
    """
    def get_last_acceptable_date(
        self,
        months_delta: int
    ) -> datetime:
        """
        Defines the last month to look for news
        """
        logging.info("Defining the last month to look for news.")
        if months_delta <= 1:
            last_acceptable_date = datetime.utcnow().replace(day=1, hour=0)
        else:
            last_acceptable_date = datetime.today() - \
                relativedelta(months=months_delta-1)
        return last_acceptable_date

    def date_filter(
        self,
        date: str
    ) -> datetime:
        """
        Converts the news' date format to datetime format
        """
        logging.info("Converting the news' date format to datetime format.")
        if 'hour' in date:
            return datetime.today()
        pattern = "%b. %d, %Y" if '.' in date else "%B %d, %Y"
        filtered_date = datetime.strptime(date, pattern)
        return filtered_date

    def download_file(
        self,
        url: str,
        date: str,
        query: str
    ) -> str:
        """
        Downloads the new's image
        """
        download_file_response = {
            'success': False
        }
        try:
            logging.info("Downloading the new's image.")
            file_name = f'{quote(query)}-{date}-{uuid.uuid4().hex}.png'
            path = os.path.join('.', 'output', file_name)

            browser = Selenium()
            browser.open_browser(
                url=url,
                service_log_path=os.path.devnull
            )
            browser.screenshot(filename=path)
            browser.close_browser()

            file_size = os.path.getsize(path)
            if file_size:
                logging.info("Successfully downloaded the new's image.")

            download_file_response.update({
                'success': True,
                'file_size': file_size,
                'picture_filename': file_name
            })

        finally:
            return download_file_response

    def build_sheet(
        self,
        extracted_data: list,
        sheet_name: str
    ) -> str:
        """
        Builds a sheet with the extracted data
        """
        logging.info("Building sheet with the extracted data.")
        header = [
            'picture_filename',
            'title',
            'description',
            'date',
            'search_phrase_count',
            'contains_money'
        ]
        header_len = len(header)
        header_cells = []

        for index in range(header_len):
            header_cells.append(string.ascii_uppercase[index]+'1')
        cell_number = str.maketrans('', '', string.digits)

        logging.info("Extracted data: %s", extracted_data)
        wb = Workbook()
        ws = wb.active

        fill_style = PatternFill(
            start_color='205792',
            end_color='205792',
            fill_type='solid'
        )

        font_style = Font(
            bold=True,
            size=12,
            color='FFFFFF',
            name='Calibri'
        )

        for index, cell in enumerate(header_cells):
            column = cell.translate(cell_number)
            ws[cell].fill = fill_style
            ws[cell].font = font_style
            ws[cell] = header[index]
            ws.column_dimensions[column].width = len(header[index]) + 25

            for index_row, process in enumerate(extracted_data):
                ws.cell(
                    column=index+1,
                    row=index_row+2,
                    value=process[header[index]]
                )

        sheet_path = os.path.join('.', 'output', sheet_name)
        wb.save(sheet_path)
        return sheet_path
