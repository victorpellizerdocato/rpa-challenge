import os
import random
import requests
from string import digits
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from dateutil.relativedelta import relativedelta


class DataHandling():
    def get_last_acceptable_date(
        months_delta
    ):
        print("Defining the last month to look for news.")
        now = datetime.utcnow()
        if months_delta <= 1:
            return now
        last_acceptable_date = datetime.today() - relativedelta(months=months_delta-1)
        return last_acceptable_date

    def date_filter(
        self,
        date
    ):
        print("Converting the news' date format to datetime format.")
        months = {
            'Jan': 1,
            'Feb': 2,
            'March': 3,
            'April': 4,
            'May': 5,
            'June': 6,
            'July': 7,
            'Aug': 8,
            'Sept': 9,
            'Oct': 10,
            'Nov': 11,
            'Dec': 12
        }
        split_date = date.split(' ')
        month = months[split_date[0].replace('.','')] # the month sometimes is misspelled
        day = split_date[1].replace(',','')
        filtered_date = datetime.strptime(f"{day}/{month}/{split_date[2]}", "%d/%m/%Y")
        return filtered_date

    def download_file(
        self,
        url
    ):
        download_response = requests.get(
            url=url,
        )
        if download_response and download_response.status_code != 200:
            raise Exception('Download failure')

        image_path = f'./images/image-{round(random.random() * (9999) + 1)}.png'
        open(image_path, "wb").write(download_response.content)
        print('Image downloaded sucessfully.')

        return os.path.abspath(image_path)

    def build_sheet(
        extracted_data,
        file_name,
        file_directory_local,
    ):
        header = [
            'picture_filename',
            'title',
            'description',
            'date',
            'search_phrase_count',
            'contains_money'
        ]
        header_len = len(header)
        header_cells = []

        for index in range(header_len):
            header_cells.append(f'{chr(index+65)}1')
        cell_number = str.maketrans('', '', digits)

        print(extracted_data)
        wb = Workbook()
        ws = wb.active

        fill_style = PatternFill(
            start_color='205792',
            end_color='205792',
            fill_type='solid'
        )

        font_style = Font(
            bold=True,
            size=12,
            color='FFFFFF',
            name='Calibri'
        )

        for index, cell in enumerate(header_cells):
            column = cell.translate(cell_number)
            ws[cell].fill = fill_style
            ws[cell].font = font_style
            ws[cell] = header[index]
            ws.column_dimensions[column].width = len(header[index]) + 25

            for index_row, process in enumerate(extracted_data):
                ws.cell(
                    column=index+1,
                    row=index_row+2,
                    value=process[header[index]]
                )

        if (not os.path.exists(file_directory_local)):
            os.makedirs(file_directory_local)
        file_directory = f'{file_directory_local}{file_name}'
        wb.save(file_directory)
        return file_directory
